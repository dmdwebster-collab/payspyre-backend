"""Turnkey-parity XLSX report exports (live test DB).

Covers ``app/services/report_exports.py`` + the admin and clinic endpoints:

  * LEDGER-sourced rows (platform_loan_transactions): payment allocations,
    reversal pairs routing to the Reversed report, full transaction dump,
  * the FALLBACK payment-split replay for ledger-less loans
    (principal-before-interest, waived skipped, overpayment capped),
  * borrower-data / originations / scheduled-payments content, opened back up
    with openpyxl from bytes,
  * vendor scoping (admin vendor_id filter; clinic principal hard-scope),
  * unknown report key -> 404, inverted window -> 422.

Run JUST this file:
    source .venv/bin/activate && \
        python -m pytest tests/test_report_exports.py -x -q
"""
from __future__ import annotations

import uuid
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from types import SimpleNamespace

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.clinic.v1.deps import ClinicPrincipal, get_current_clinic_user
from app.api.clinic.v1.router import clinic_router
from app.api.v1.api import api_router
from app.core.auth import get_current_user
from app.db.base import get_db
from app.models.loan import Vendor
from app.models.platform.credit_application import PlatformCreditApplication
from app.models.platform.credit_product import PlatformCreditProduct
from app.models.platform.loan import (
    PlatformLoan,
    PlatformLoanCustomTransaction,
    PlatformLoanPayment,
    PlatformLoanScheduleItem,
    PlatformLoanTransaction,
)
from app.models.platform.patient import PlatformPatient
from app.services.report_exports import (
    REPORT_KEYS,
    allocate_payment_splits,
    generate_report,
)

_ADMIN_BASE = "/api/v1/admin/reports"
_CLINIC_BASE = "/api/clinic/v1/reports"


# ---------------------------------------------------------------------------
# Seed helpers
# ---------------------------------------------------------------------------


def _mk_vendor(db, name="Clinic A", dba=None) -> Vendor:
    v = Vendor(
        business_name=name,
        dba_name=dba,
        business_type="corporation",
        contact_name="Owner",
        email=f"{uuid.uuid4().hex}@example.com",
        phone="+15555550100",
        address_line1="1 Main St",
        city="Kelowna",
        province="BC",
        postal_code="V1Y1Y1",
        status="active",
    )
    db.add(v)
    db.flush()
    return v


def _product(db) -> PlatformCreditProduct:
    return (
        db.query(PlatformCreditProduct)
        .filter(PlatformCreditProduct.code == "dental_full_arch_v1")
        .first()
    )


def _mk_loan(
    db,
    vendor,
    *,
    first="Jordan",
    last="Lee",
    principal=120_000,
    rate_bps=1200,
    term=2,
    disbursed_at=None,
) -> PlatformLoan:
    patient = PlatformPatient(legal_first_name=first, legal_last_name=last)
    db.add(patient)
    db.flush()
    app = PlatformCreditApplication(
        patient_id=patient.id,
        credit_product_id=_product(db).id,
        credit_product_version=1,
        requested_amount_cents=principal,
        requested_amount_source="clinic",
        vendor_id=vendor.id,
        status="approved",
        first_name=first,
        last_name=last,
        email=f"{first.lower()}.{last.lower()}@example.com",
        main_phone="+16045550111",
        residence_street="12 Elm St",
        residence_city="Kelowna",
        residence_province="BC",
        residence_postal_code="V1Y2B2",
    )
    db.add(app)
    db.flush()
    loan = PlatformLoan(
        application_id=app.id,
        principal_cents=principal,
        annual_rate_bps=rate_bps,
        term_months=term,
        status="active",
        principal_balance_cents=principal,
        disbursement_status="completed",
        disbursed_at=disbursed_at or datetime(2026, 6, 1, 12, tzinfo=timezone.utc),
    )
    db.add(loan)
    db.flush()
    return loan


def _mk_schedule(db, loan, items):
    """items: list of (n, due, principal, interest)."""
    for n, due, principal, interest in items:
        db.add(
            PlatformLoanScheduleItem(
                loan_id=loan.id,
                installment_number=n,
                due_date=due,
                principal_cents=principal,
                interest_cents=interest,
                total_cents=principal + interest,
                status="scheduled",
                paid_cents=0,
            )
        )
    db.flush()


def _mk_payment(db, loan, amount, received_at, ref=None) -> PlatformLoanPayment:
    p = PlatformLoanPayment(
        loan_id=loan.id,
        amount_cents=amount,
        received_at=received_at,
        method="zumrails",
        external_ref=ref,
    )
    db.add(p)
    db.flush()
    return p


def _mk_txn(
    db,
    loan,
    seq,
    txn_type,
    amount,
    *,
    principal=0,
    interest=0,
    fees=0,
    effective=None,
    payment_type=None,
    reverses=None,
) -> PlatformLoanTransaction:
    t = PlatformLoanTransaction(
        loan_id=loan.id,
        seq=seq,
        reference=f"V-{loan.id}-{seq}",
        txn_type=txn_type,
        payment_type=payment_type,
        amount_cents=amount,
        principal_cents=principal,
        interest_cents=interest,
        fees_cents=fees,
        effective_date=effective or date(2026, 6, 15),
        processing_date=effective or date(2026, 6, 15),
        reverses_transaction_id=reverses,
        created_by="test",
    )
    db.add(t)
    db.flush()
    return t


def _sheet_rows(content: bytes) -> list[list]:
    ws = load_workbook(BytesIO(content)).active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _data_rows(content: bytes) -> list[list]:
    """Rows below the header row (row 8), minus the totals row (formula strings)."""
    rows = _sheet_rows(content)[8:]
    return [r for r in rows if r and not any(isinstance(c, str) and c.startswith("=") for c in r)]


# ---------------------------------------------------------------------------
# Pure: fallback payment split replay
# ---------------------------------------------------------------------------


def test_allocate_payment_splits_principal_before_interest():
    sched = [
        SimpleNamespace(id=1, installment_number=1, principal_cents=800, interest_cents=200,
                        total_cents=1000, status="scheduled"),
        SimpleNamespace(id=2, installment_number=2, principal_cents=900, interest_cents=100,
                        total_cents=1000, status="scheduled"),
    ]
    t0 = datetime(2026, 6, 10, tzinfo=timezone.utc)
    pays = [
        SimpleNamespace(id="p1", amount_cents=500, received_at=t0, created_at=t0),
        SimpleNamespace(id="p2", amount_cents=1000, received_at=t0 + timedelta(days=1),
                        created_at=t0 + timedelta(days=1)),
    ]
    splits = allocate_payment_splits(sched, pays)
    # p1: 500 into installment 1 -> all principal (800 outstanding principal).
    assert splits["p1"] == {"principal_cents": 500, "interest_cents": 0, "applied_cents": 500}
    # p2: 500 finishes installment 1 (300 principal + 200 interest),
    #     then 500 into installment 2 -> all principal.
    assert splits["p2"] == {"principal_cents": 800, "interest_cents": 200, "applied_cents": 1000}


def test_allocate_payment_splits_skips_waived_and_caps_overpayment():
    sched = [
        SimpleNamespace(id=1, installment_number=1, principal_cents=800, interest_cents=200,
                        total_cents=1000, status="waived"),
        SimpleNamespace(id=2, installment_number=2, principal_cents=900, interest_cents=100,
                        total_cents=1000, status="scheduled"),
    ]
    t0 = datetime(2026, 6, 10, tzinfo=timezone.utc)
    pays = [SimpleNamespace(id="p1", amount_cents=1500, received_at=t0, created_at=t0)]
    splits = allocate_payment_splits(sched, pays)
    # Waived installment consumes nothing; overpayment beyond schedule stays
    # undistributed (applied < amount).
    assert splits["p1"] == {"principal_cents": 900, "interest_cents": 100, "applied_cents": 1000}


# ---------------------------------------------------------------------------
# Service: loan-level reports
# ---------------------------------------------------------------------------


def test_borrower_data_report_content(db_session):
    vendor = _mk_vendor(db_session, name="New Teeth Today", dba="NTT Kelowna")
    loan = _mk_loan(db_session, vendor, first="Ava", last="Stone", principal=240_000, term=2)
    _mk_schedule(
        db_session,
        loan,
        [(1, date(2026, 7, 1), 120_000, 2_000), (2, date(2026, 8, 1), 120_000, 1_000)],
    )
    db_session.commit()

    result = generate_report(db_session, "borrower-data")
    assert result.filename.startswith("payspyre_borrower_data_")
    rows = _sheet_rows(result.content)
    assert rows[0][0] == "PaySpyre Financial Inc."
    assert rows[7][0] == "Vendor"  # header row is row 8
    data = _data_rows(result.content)
    row = next(r for r in data if r[3] == "Ava Stone")
    assert row[0] == "New Teeth Today"
    assert row[1] == "NTT Kelowna"
    assert row[4] == "ava.stone@example.com"
    assert row[8] == "BC"
    assert row[10] == "Active"
    assert row[13] == 2400.00           # Loan Amount in dollars
    assert row[14] == pytest.approx(1215.00)  # avg installment (243000/2 cents)
    assert row[15] == pytest.approx(0.12)     # 1200 bps as fraction
    assert row[16] == "2 months"
    assert row[17] == 2400.00           # outstanding principal


def test_originations_window_filters(db_session):
    vendor = _mk_vendor(db_session)
    _mk_loan(db_session, vendor, first="In", last="Window",
             disbursed_at=datetime(2026, 6, 15, tzinfo=timezone.utc))
    _mk_loan(db_session, vendor, first="Out", last="Side",
             disbursed_at=datetime(2026, 3, 1, tzinfo=timezone.utc))
    db_session.commit()

    result = generate_report(
        db_session, "loan-originations",
        date_from=date(2026, 6, 1), date_to=date(2026, 6, 30),
    )
    names = [r[3] for r in _data_rows(result.content)]
    assert "In Window" in names
    assert "Out Side" not in names


# ---------------------------------------------------------------------------
# Service: LEDGER-sourced payment/transaction reports
# ---------------------------------------------------------------------------


def test_payments_received_from_ledger_excludes_reversed(db_session):
    vendor = _mk_vendor(db_session)
    loan = _mk_loan(db_session, vendor, first="Led", last="Ger", principal=5_000)
    keep = _mk_txn(db_session, loan, 1, "payment", 1000, principal=800, interest=150,
                   fees=50, effective=date(2026, 6, 10), payment_type="eft")
    gone = _mk_txn(db_session, loan, 2, "payment", 700, principal=700,
                   effective=date(2026, 6, 12), payment_type="cash")
    _mk_txn(db_session, loan, 3, "reversal", 700, effective=date(2026, 6, 14),
            reverses=gone.id)
    db_session.commit()

    result = generate_report(db_session, "payments-received", vendor_id=vendor.id)
    data = _data_rows(result.content)
    assert len(data) == 1  # the reversed payment routes to the Reversed report
    row = data[0]
    assert row[5] == datetime(2026, 6, 10)  # payment (effective) date
    assert row[6] == 10.00                  # amount
    assert row[7] == 0.50                   # fees bucket -> Admin Fees Pd
    assert row[9] == 0.50                   # Fee(s) Paid
    assert row[10] == 1.50                  # interest
    assert row[11] == 8.00                  # principal
    assert row[13] == "Bank Transfer"       # TL payment-type vocabulary
    assert row[14] == keep.reference


def test_payments_reversed_report_shows_original_payment(db_session):
    vendor = _mk_vendor(db_session)
    loan = _mk_loan(db_session, vendor, first="Rev", last="Ersed", principal=5_000)
    original = _mk_txn(db_session, loan, 1, "payment", 900, principal=800, interest=100,
                       effective=date(2026, 6, 10), payment_type="eft")
    _mk_txn(db_session, loan, 2, "reversal", 900, effective=date(2026, 6, 20),
            reverses=original.id)
    db_session.commit()

    result = generate_report(db_session, "payments-reversed", vendor_id=vendor.id)
    data = _data_rows(result.content)
    assert len(data) == 1
    row = data[0]
    assert row[4] == datetime(2026, 6, 10)  # ORIGINAL payment date
    assert row[5] == 9.00
    assert row[9] == 1.00                   # interest
    assert row[10] == 8.00                  # principal
    assert row[13] == original.reference

    # Window on the reversal date: a window covering only the payment date
    # (not the reversal) excludes the row.
    result2 = generate_report(
        db_session, "payments-reversed", vendor_id=vendor.id,
        date_from=date(2026, 6, 9), date_to=date(2026, 6, 11),
    )
    assert _data_rows(result2.content) == []


def test_loan_transactions_ledger_dump(db_session):
    vendor = _mk_vendor(db_session)
    loan = _mk_loan(db_session, vendor, first="Tx", last="Ledger", principal=5_000)
    _mk_txn(db_session, loan, 1, "disbursement", 5000, effective=date(2026, 6, 1))
    _mk_txn(db_session, loan, 2, "payment", 1000, principal=900, interest=100,
            effective=date(2026, 6, 15), payment_type="eft")
    db_session.commit()

    result = generate_report(db_session, "loan-transactions", vendor_id=vendor.id)
    data = _data_rows(result.content)
    assert [r[5] for r in data] == ["Disbursement", "Payment"]
    payment_row = data[1]
    assert payment_row[6] == "Bank Transfer"
    assert payment_row[7] == 10.00   # amount
    assert payment_row[8] == 1.00    # interest
    assert payment_row[9] == 9.00    # principal
    assert payment_row[12].endswith("-2")  # Dave's {vendor}-{loan}-{seq} reference


# ---------------------------------------------------------------------------
# Service: FALLBACK path (loans with no ledger rows)
# ---------------------------------------------------------------------------


def test_payments_received_fallback_split_and_scoping(db_session):
    vendor_a = _mk_vendor(db_session, name="Vendor A")
    vendor_b = _mk_vendor(db_session, name="Vendor B")
    loan_a = _mk_loan(db_session, vendor_a, first="Ann", last="Ay", principal=2_000)
    _mk_schedule(db_session, loan_a, [(1, date(2026, 7, 1), 800, 200),
                                      (2, date(2026, 8, 1), 900, 100)])
    _mk_payment(db_session, loan_a, 500, datetime(2026, 6, 10, tzinfo=timezone.utc), ref="zr-1")
    _mk_payment(db_session, loan_a, 1000, datetime(2026, 6, 11, tzinfo=timezone.utc), ref="zr-2")
    loan_b = _mk_loan(db_session, vendor_b, first="Bob", last="Bee", principal=2_000)
    _mk_schedule(db_session, loan_b, [(1, date(2026, 7, 1), 900, 100)])
    _mk_payment(db_session, loan_b, 500, datetime(2026, 6, 12, tzinfo=timezone.utc), ref="zr-3")
    db_session.commit()

    result = generate_report(db_session, "payments-received", vendor_id=vendor_a.id)
    data = _data_rows(result.content)
    assert [r[3] for r in data] == ["Ann Ay", "Ann Ay"]  # vendor B excluded
    first, second = data
    # p1: all principal; p2: 300 principal to finish inst 1 + 200 interest + 500 principal.
    assert first[10] == 0.0 and first[11] == 5.00
    assert second[10] == 2.00 and second[11] == 8.00
    assert first[13] == "Bank Transfer (Zumrails)"
    assert first[14] == "zr-1"


def test_failed_report_is_headers_only(db_session):
    result = generate_report(db_session, "payments-failed")
    assert result.row_count == 0
    assert _sheet_rows(result.content)[7][0] == "Vendor"


def test_loan_transactions_fallback_synthesis(db_session):
    vendor = _mk_vendor(db_session)
    loan = _mk_loan(db_session, vendor, first="Tx", last="Case", principal=1_000)
    _mk_schedule(db_session, loan, [(1, date(2026, 7, 1), 900, 100)])
    _mk_payment(db_session, loan, 1000, datetime(2026, 6, 20, tzinfo=timezone.utc), ref="zr-9")
    db_session.commit()

    result = generate_report(db_session, "loan-transactions", vendor_id=vendor.id)
    data = _data_rows(result.content)
    types = [r[5] for r in data]
    assert types == ["Disbursement", "Payment"]
    payment_row = data[1]
    assert payment_row[7] == 10.00   # amount
    assert payment_row[8] == 1.00    # interest
    assert payment_row[9] == 9.00    # principal
    assert payment_row[12] == "zr-9"


def test_scheduled_payments_statuses_and_custom_txn(db_session):
    vendor = _mk_vendor(db_session)
    loan = _mk_loan(db_session, vendor, first="Sked", last="Yule", principal=3_000)
    today = datetime.now(timezone.utc).date()
    _mk_schedule(
        db_session,
        loan,
        [
            (1, today - timedelta(days=40), 900, 100),  # far past due
            (2, today - timedelta(days=5), 900, 100),   # inside grace (15d)
            (3, today + timedelta(days=25), 900, 100),  # future
        ],
    )
    db_session.add(
        PlatformLoanCustomTransaction(
            loan_id=loan.id,
            scheduled_date=today + timedelta(days=10),
            amount_cents=1_500,
            repayment_mode="regular",
            status="scheduled",
            comment="Borrower request, authorization obtained",
            created_by="test",
        )
    )
    db_session.commit()

    result = generate_report(db_session, "scheduled-payments", vendor_id=vendor.id)
    data = _data_rows(result.content)
    assert [r[6] for r in data] == ["Past Due", "Grace", "Scheduled (Custom)", "Scheduled"]
    assert data[2][5] == 15.00  # the custom transaction's amount


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


def _admin():
    return SimpleNamespace(
        id=uuid.uuid4(),
        roles=[SimpleNamespace(role=SimpleNamespace(name="admin"))],
    )


@pytest.fixture
def admin_client(db_session):
    app = FastAPI()
    app.include_router(api_router, prefix="/api/v1")
    app.dependency_overrides[get_db] = lambda: db_session
    app.dependency_overrides[get_current_user] = _admin
    yield TestClient(app)
    app.dependency_overrides.clear()


def test_admin_endpoint_returns_xlsx(admin_client, db_session):
    vendor = _mk_vendor(db_session)
    _mk_loan(db_session, vendor, first="End", last="Point")
    db_session.commit()

    r = admin_client.get(f"{_ADMIN_BASE}/exports/borrower-data")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="payspyre_borrower_data_' in r.headers["content-disposition"]
    assert int(r.headers["x-report-rows"]) >= 1
    names = [row[3] for row in _data_rows(r.content)]
    assert "End Point" in names


def test_admin_endpoint_catalog_and_errors(admin_client):
    r = admin_client.get(f"{_ADMIN_BASE}/exports")
    assert r.status_code == 200
    assert {e["key"] for e in r.json()} == set(REPORT_KEYS)

    assert admin_client.get(f"{_ADMIN_BASE}/exports/nope").status_code == 404
    r = admin_client.get(
        f"{_ADMIN_BASE}/exports/borrower-data",
        params={"date_from": "2026-07-01", "date_to": "2026-06-01"},
    )
    assert r.status_code == 422


def test_clinic_endpoint_hard_scopes_to_own_vendor(db_session):
    vendor_a = _mk_vendor(db_session, name="Mine")
    vendor_b = _mk_vendor(db_session, name="Theirs")
    _mk_loan(db_session, vendor_a, first="My", last="Patient")
    _mk_loan(db_session, vendor_b, first="Their", last="Patient")
    db_session.commit()

    app = FastAPI()
    app.include_router(clinic_router, prefix="/api/clinic/v1")
    app.dependency_overrides[get_db] = lambda: db_session
    app.dependency_overrides[get_current_clinic_user] = lambda: ClinicPrincipal(
        user=SimpleNamespace(id=uuid.uuid4()), vendor_id=vendor_a.id
    )
    client = TestClient(app)

    r = client.get(f"{_CLINIC_BASE}/exports/borrower-data")
    assert r.status_code == 200, r.text
    names = [row[3] for row in _data_rows(r.content)]
    assert "My Patient" in names
    assert "Their Patient" not in names
    app.dependency_overrides.clear()

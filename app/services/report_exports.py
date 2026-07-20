"""Turnkey-parity operational report exports (XLSX).

Replicates the seven PaySpyre-branded reports Dave built inside Turnkey Lender
with Smart Marker templates (2025-08-18 template export):

  * borrower-data        Borrower Data
  * loan-originations    New Loan Originations
  * payments-received    Successful Loan Payments
  * payments-failed      Failed Loan Payments
  * payments-reversed    Reversed Loan Payments
  * loan-transactions    Loan Transactions
  * scheduled-payments   Scheduled Payments

Column layouts mirror the TL templates 1:1 so vendors and ops receive the same
artifact after cutover. docs/turnkey_parity/report_exports.md maps every column
to its platform source and records the known gaps.

DATA SOURCES. The immutable money ledger (``platform_loan_transactions``,
WS-A / migration 049) is the source of truth for payment and transaction rows:
it stores each row's principal / interest / fees allocation, Dave's
``{vendor}-{loan}-{seq}`` reference, the dual effective/processing dates, and
compensating ``reversal`` rows — which is what makes the Reversed report real.
Loans with NO ledger rows (a pre-cutover book that never got its
reconciliation row) fall back to replaying ``platform_loan_payments`` against
the schedule with exactly the allocation order of the pre-ledger
``record_payment`` (principal before interest, oldest installment first).

Money is integer cents in the DB; cells are written in dollars with a currency
number format.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from io import BytesIO
from typing import Callable, Iterable, Optional, Sequence
from uuid import UUID

from sqlalchemy.orm import Session, selectinload

from app.core.config import settings
from app.models.loan import Vendor
from app.models.platform.credit_application import PlatformCreditApplication
from app.models.platform.loan import (
    PlatformLoan,
    PlatformLoanPayment,
    PlatformLoanScheduleItem,
    PlatformLoanTransaction,
)
from app.models.platform.patient import PlatformPatient

# Branding block — rows 1..5 of every TL template.
COMPANY_HEADER = (
    "PaySpyre Financial Inc.",
    "100-2033 Gordon Dr.",
    "Kelowna, BC. V1Y 3J2",
    "1 (877) 729 1973",
    "vendor@payspyre.com",
)
COMPANY_WEBSITE = "www.payspyre.com"

_LOAN_STATUS_LABELS = {
    "pending_disbursement": "Pending Disbursement",
    "active": "Active",
    "paid_off": "Paid Off",
    "delinquent": "Delinquent",
    "charged_off": "Charged Off",
    "cancelled": "Cancelled",
}

# Ledger payment_type -> TL's payment-type vocabulary (Cash; Check; Bank
# Transfer; Card).
_PAYMENT_TYPE_LABELS = {
    "cash": "Cash",
    "check": "Check",
    "eft": "Bank Transfer",
    "credit_card": "Card",
    "adjustment": "Adjustment",
}

_TXN_TYPE_LABELS = {
    "payment": "Payment",
    "disbursement": "Disbursement",
    "fee": "Fee",
    "adjustment": "Adjustment",
    "reversal": "Reversal",
}


# Cell kinds drive the openpyxl number format per column.
#   text | money | percent | int | date | datetime
@dataclass(frozen=True)
class Col:
    header: str
    kind: str = "text"
    total: bool = False  # sum this column into the totals row (money only)


@dataclass(frozen=True)
class ReportResult:
    key: str
    title: str
    filename: str
    content: bytes
    row_count: int


# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------


def _dollars(cents: Optional[int]) -> Optional[float]:
    return None if cents is None else round(cents / 100.0, 2)


def _loan_display_id(loan: PlatformLoan) -> str:
    """Human-facing loan id: the legacy TL account number for migrated loans
    (so vendors keep the numbers they know), else the platform UUID."""
    return loan.legacy_account_number or str(loan.id)


def _full_name(application, patient) -> str:
    if application is not None:
        name = " ".join(
            p for p in (application.first_name, application.last_name) if p
        ).strip()
        if name:
            return name
    if patient is not None:
        name = " ".join(
            p
            for p in (patient.legal_first_name, patient.legal_last_name)
            if p
        ).strip()
        if name:
            return name
    return ""


def _vendor_names(vendor: Optional[Vendor]) -> tuple[str, str]:
    """(Vendor, Provider) columns.

    TL models Vendor -> Store(Provider); the platform has a single ``vendors``
    entity, so Vendor = business_name and Provider = dba_name (falling back to
    business_name). When a provider/store entity lands, swap the second value.
    """
    if vendor is None:
        return "", ""
    return vendor.business_name or "", vendor.dba_name or vendor.business_name or ""


def _method_label(method: Optional[str]) -> str:
    """Prettify a raw ``platform_loan_payments.method`` (fallback path only)."""
    if not method:
        return ""
    return {"zumrails": "Bank Transfer (Zumrails)"}.get(
        method, method.replace("_", " ").title()
    )


def _close_date(loan: PlatformLoan) -> Optional[date]:
    """TL Close Date ~ when the loan left the live book. The platform doesn't
    store a dedicated close timestamp; the last status update is the closest
    proxy for terminal loans (documented gap)."""
    if loan.status in ("paid_off", "charged_off", "cancelled") and loan.updated_at:
        return loan.updated_at.date()
    return None


def _average_installment_cents(schedule: Sequence[PlatformLoanScheduleItem]) -> Optional[int]:
    if not schedule:
        return None
    return round(sum(s.total_cents for s in schedule) / len(schedule))


def allocate_payment_splits(
    schedule: Sequence[PlatformLoanScheduleItem],
    payments: Sequence[PlatformLoanPayment],
) -> dict:
    """FALLBACK for ledger-less loans: derive each payment's (principal,
    interest) split by replaying the full payment history oldest-first against
    the schedule.

    Mirrors the pre-ledger ``record_payment`` allocation: payments fill the
    oldest unpaid installment first; within an installment, principal fills
    before interest. Waived installments are skipped. Any amount beyond the
    schedule total is left undistributed (visible as ``applied`` < amount).

    Returns {payment.id: {"principal_cents": int, "interest_cents": int,
                          "applied_cents": int}}.
    """
    items = sorted(schedule, key=lambda s: s.installment_number)
    # Replay from zero — DB paid_cents already includes every payment.
    filled = {s.id: 0 for s in items}

    out: dict = {}
    ordered = sorted(payments, key=lambda p: (p.received_at, p.created_at or p.received_at))
    for payment in ordered:
        remaining = payment.amount_cents
        principal_total = 0
        interest_total = 0
        for item in items:
            if remaining <= 0:
                break
            if item.status == "waived":
                continue
            outstanding = item.total_cents - filled[item.id]
            if outstanding <= 0:
                continue
            applied = min(remaining, outstanding)
            principal_outstanding = max(0, item.principal_cents - filled[item.id])
            principal_applied = min(applied, principal_outstanding)

            filled[item.id] += applied
            remaining -= applied
            principal_total += principal_applied
            interest_total += applied - principal_applied

        out[payment.id] = {
            "principal_cents": principal_total,
            "interest_cents": interest_total,
            "applied_cents": payment.amount_cents - remaining,
        }
    return out


def _expected_status(item: PlatformLoanScheduleItem, as_of: date) -> str:
    """TL expected-payment status vocabulary: Scheduled / Grace / Past Due,
    with Partial preserved for partially-covered installments."""
    if item.due_date >= as_of:
        return "Partial" if item.status == "partial" else "Scheduled"
    days_over = (as_of - item.due_date).days
    if days_over <= settings.GRACE_DAYS:
        return "Grace"
    return "Past Due"


def _ledger_rows(loan: PlatformLoan) -> list[PlatformLoanTransaction]:
    return sorted(loan.transactions, key=lambda t: t.seq)


def _reversed_txn_ids(txns: Iterable[PlatformLoanTransaction]) -> set:
    return {
        t.reverses_transaction_id
        for t in txns
        if t.txn_type == "reversal" and t.reverses_transaction_id is not None
    }


# ---------------------------------------------------------------------------
# Query plumbing
# ---------------------------------------------------------------------------


def _loan_rows(
    db: Session,
    vendor_id: Optional[UUID],
) -> list[tuple[PlatformLoan, PlatformCreditApplication, Vendor, PlatformPatient]]:
    """(loan, application, vendor, patient) tuples, vendor-scoped when asked.

    Joins are OUTER so migrated loans without an application still export in
    the all-vendors view; a vendor_id filter implies the application join and
    therefore naturally excludes application-less loans (they belong to no
    vendor). Collections (schedule / payments / ledger / custom txns) are
    selectin-loaded — one extra query each instead of per-loan N+1.
    """
    q = (
        db.query(PlatformLoan, PlatformCreditApplication, Vendor, PlatformPatient)
        .outerjoin(
            PlatformCreditApplication,
            PlatformLoan.application_id == PlatformCreditApplication.id,
        )
        .outerjoin(Vendor, PlatformCreditApplication.vendor_id == Vendor.id)
        .outerjoin(
            PlatformPatient, PlatformCreditApplication.patient_id == PlatformPatient.id
        )
        .options(
            selectinload(PlatformLoan.schedule),
            selectinload(PlatformLoan.payments),
            selectinload(PlatformLoan.transactions),
            selectinload(PlatformLoan.custom_transactions),
        )
    )
    if vendor_id is not None:
        q = q.filter(PlatformCreditApplication.vendor_id == vendor_id)
    return q.order_by(PlatformLoan.created_at.asc()).all()


def _window_bounds(
    date_from: Optional[date], date_to: Optional[date]
) -> tuple[Optional[datetime], Optional[datetime]]:
    """Inclusive date window -> tz-aware datetime bounds [from 00:00, to 24:00)."""
    frm = (
        datetime.combine(date_from, time.min, tzinfo=timezone.utc)
        if date_from
        else None
    )
    to = (
        datetime.combine(date_to, time.max, tzinfo=timezone.utc) if date_to else None
    )
    return frm, to


def _in_window(value: Optional[datetime], frm, to) -> bool:
    if value is None:
        return False
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    if frm is not None and value < frm:
        return False
    if to is not None and value > to:
        return False
    return True


def _date_in_window(value: Optional[date], date_from, date_to) -> bool:
    if value is None:
        return False
    if date_from is not None and value < date_from:
        return False
    if date_to is not None and value > date_to:
        return False
    return True


# ---------------------------------------------------------------------------
# Report builders — each returns (columns, rows)
# ---------------------------------------------------------------------------


def _build_borrower_data(db, vendor_id, date_from, date_to):
    cols = [
        Col("Vendor"),
        Col("Provider"),
        Col("Loan ID"),
        Col("Full Name"),
        Col("Email"),
        Col("Phone"),
        Col("Street"),
        Col("City"),
        Col("Province"),
        Col("Postal Code"),
        Col("Status"),
        Col("Activation Date", "date"),
        Col("Close Date", "date"),
        Col("Loan Amount", "money", total=True),
        Col("Average Installment", "money"),
        Col("Interest Rate", "percent"),
        Col("Loan Term"),
        Col("Outstanding Principal", "money", total=True),
    ]
    frm, to = _window_bounds(date_from, date_to)
    rows = []
    for loan, app, vendor, patient in _loan_rows(db, vendor_id):
        origin = loan.disbursed_at or loan.created_at
        if (date_from or date_to) and not _in_window(origin, frm, to):
            continue
        vend, prov = _vendor_names(vendor)
        rows.append(
            [
                vend,
                prov,
                _loan_display_id(loan),
                _full_name(app, patient),
                (app.email if app else None) or (patient.email if patient else "") or "",
                (app.main_phone if app else None)
                or (patient.phone_e164 if patient else "")
                or "",
                (app.residence_street if app else "") or "",
                (app.residence_city if app else "") or "",
                (app.residence_province if app else "") or "",
                (app.residence_postal_code if app else "") or "",
                _LOAN_STATUS_LABELS.get(loan.status, loan.status),
                loan.disbursed_at.date() if loan.disbursed_at else None,
                _close_date(loan),
                _dollars(loan.principal_cents),
                _dollars(_average_installment_cents(loan.schedule)),
                loan.annual_rate_bps / 10_000,
                f"{loan.term_months} months",
                _dollars(loan.principal_balance_cents),
            ]
        )
    return cols, rows


def _build_loan_originations(db, vendor_id, date_from, date_to):
    cols = [
        Col("Vendor"),
        Col("Provider"),
        Col("Loan ID"),
        Col("Full Name"),
        Col("Activation Date", "date"),
        Col("Loan Amount", "money", total=True),
        Col("Average Installment", "money"),
        Col("Interest Rate", "percent"),
        Col("Loan Term"),
        Col("Status"),
    ]
    frm, to = _window_bounds(date_from, date_to)
    rows = []
    for loan, app, vendor, patient in _loan_rows(db, vendor_id):
        if loan.disbursed_at is None:
            continue
        if (date_from or date_to) and not _in_window(loan.disbursed_at, frm, to):
            continue
        vend, prov = _vendor_names(vendor)
        rows.append(
            [
                vend,
                prov,
                _loan_display_id(loan),
                _full_name(app, patient),
                loan.disbursed_at.date(),
                _dollars(loan.principal_cents),
                _dollars(_average_installment_cents(loan.schedule)),
                loan.annual_rate_bps / 10_000,
                f"{loan.term_months} months",
                _LOAN_STATUS_LABELS.get(loan.status, loan.status),
            ]
        )
    return cols, rows


_PAYMENT_COLS = [
    Col("Vendor"),
    Col("Provider"),
    Col("Loan ID"),
    Col("Full Name"),
    Col("Actual Time", "datetime"),
    Col("Payment Date", "date"),
    Col("Payment Amount", "money", total=True),
    Col("Admin Fees Pd", "money", total=True),
    Col("Penalty Fees Pd", "money", total=True),
    Col("Fee(s) Paid", "money", total=True),
    Col("Interest Paid", "money", total=True),
    Col("Principal Paid", "money", total=True),
    Col("*Principal Balance", "money"),
    Col("Payment Method"),
    Col("Reference"),
]


def _build_payments_received(db, vendor_id, date_from, date_to):
    frm, to = _window_bounds(date_from, date_to)
    rows = []
    for loan, app, vendor, patient in _loan_rows(db, vendor_id):
        vend, prov = _vendor_names(vendor)
        name = _full_name(app, patient)
        loan_id = _loan_display_id(loan)
        balance = _dollars(loan.principal_balance_cents)

        txns = _ledger_rows(loan)
        if txns:
            # LEDGER path: real allocations; a payment that has a compensating
            # reversal row belongs to the Reversed report, not this one.
            reversed_ids = _reversed_txn_ids(txns)
            for t in txns:
                if t.txn_type != "payment" or t.id in reversed_ids:
                    continue
                if (date_from or date_to) and not _date_in_window(
                    t.effective_date, date_from, date_to
                ):
                    continue
                rows.append(
                    [
                        vend,
                        prov,
                        loan_id,
                        name,
                        t.created_at,
                        t.effective_date,
                        _dollars(t.amount_cents),
                        # The ledger has one fee bucket (no admin/penalty split
                        # yet) — carried in Admin Fees Pd; Fee(s) Paid is the
                        # TL template's admin+penalty sum.
                        _dollars(t.fees_cents),
                        0.0,
                        _dollars(t.fees_cents),
                        _dollars(t.interest_cents),
                        _dollars(t.principal_cents),
                        balance,
                        _PAYMENT_TYPE_LABELS.get(t.payment_type or "", ""),
                        t.reference,
                    ]
                )
        elif loan.payments:
            # FALLBACK path (pre-ledger book): replay-derived splits.
            splits = allocate_payment_splits(loan.schedule, loan.payments)
            for payment in loan.payments:
                if (date_from or date_to) and not _in_window(
                    payment.received_at, frm, to
                ):
                    continue
                split = splits[payment.id]
                rows.append(
                    [
                        vend,
                        prov,
                        loan_id,
                        name,
                        payment.received_at,
                        payment.received_at.date(),
                        _dollars(payment.amount_cents),
                        0.0,
                        0.0,
                        0.0,
                        _dollars(split["interest_cents"]),
                        _dollars(split["principal_cents"]),
                        balance,
                        _method_label(payment.method),
                        payment.external_ref or "",
                    ]
                )
    rows.sort(key=lambda r: (r[5], str(r[2])))
    return _PAYMENT_COLS, rows


def _build_payments_failed(db, vendor_id, date_from, date_to):
    # GAP: failed collection attempts are not persisted (the ledger records
    # money that MOVED; a failed pull moves nothing and no attempt table exists
    # yet). Headers ship now so the ops workflow and frontend contract exist;
    # rows appear when payment-attempt tracking lands (auto-collection WS).
    return _PAYMENT_COLS, []


def _build_payments_reversed(db, vendor_id, date_from, date_to):
    cols = [c for c in _PAYMENT_COLS if c.header != "Actual Time"]
    rows = []
    for loan, app, vendor, patient in _loan_rows(db, vendor_id):
        txns = _ledger_rows(loan)
        if not txns:
            continue
        by_id = {t.id: t for t in txns}
        vend, prov = _vendor_names(vendor)
        name = _full_name(app, patient)
        loan_id = _loan_display_id(loan)
        balance = _dollars(loan.principal_balance_cents)
        for t in txns:
            if t.txn_type != "reversal" or t.reverses_transaction_id is None:
                continue
            original = by_id.get(t.reverses_transaction_id)
            if original is None or original.txn_type != "payment":
                continue
            # Window on WHEN it was reversed; the row shows the original
            # payment's date and amounts (TL's reversed-payments layout).
            if (date_from or date_to) and not _date_in_window(
                t.effective_date, date_from, date_to
            ):
                continue
            rows.append(
                [
                    vend,
                    prov,
                    loan_id,
                    name,
                    original.effective_date,
                    _dollars(original.amount_cents),
                    _dollars(original.fees_cents),
                    0.0,
                    _dollars(original.fees_cents),
                    _dollars(original.interest_cents),
                    _dollars(original.principal_cents),
                    balance,
                    _PAYMENT_TYPE_LABELS.get(original.payment_type or "", ""),
                    original.reference,
                ]
            )
    rows.sort(key=lambda r: (r[4], str(r[2])))
    return cols, rows


def _build_loan_transactions(db, vendor_id, date_from, date_to):
    cols = [
        Col("Vendor"),
        Col("Provider"),
        Col("Loan ID"),
        Col("Full Name"),
        Col("Transaction Date", "date"),
        Col("Type"),
        Col("Method"),
        Col("Amount", "money", total=True),
        Col("Interest", "money", total=True),
        Col("Principal", "money", total=True),
        Col("Fee", "money", total=True),
        Col("*Outstanding Principal", "money"),
        Col("Reference"),
    ]
    frm, to = _window_bounds(date_from, date_to)
    rows = []
    for loan, app, vendor, patient in _loan_rows(db, vendor_id):
        vend, prov = _vendor_names(vendor)
        name = _full_name(app, patient)
        loan_id = _loan_display_id(loan)
        balance = _dollars(loan.principal_balance_cents)

        txns = _ledger_rows(loan)
        if txns:
            # LEDGER path: the ledger IS the transaction log — dump it.
            for t in txns:
                if (date_from or date_to) and not _date_in_window(
                    t.effective_date, date_from, date_to
                ):
                    continue
                rows.append(
                    [
                        vend,
                        prov,
                        loan_id,
                        name,
                        t.effective_date,
                        _TXN_TYPE_LABELS.get(t.txn_type, t.txn_type),
                        _PAYMENT_TYPE_LABELS.get(t.payment_type or "", ""),
                        _dollars(t.amount_cents),
                        _dollars(t.interest_cents),
                        _dollars(t.principal_cents),
                        _dollars(t.fees_cents),
                        balance,
                        t.reference,
                    ]
                )
            continue

        # FALLBACK path (pre-ledger book): synthesize from disbursed_at + the
        # payment-replay splits.
        if loan.disbursed_at is not None and (
            not (date_from or date_to) or _in_window(loan.disbursed_at, frm, to)
        ):
            rows.append(
                [
                    vend,
                    prov,
                    loan_id,
                    name,
                    loan.disbursed_at.date(),
                    "Disbursement",
                    "Bank Transfer",
                    _dollars(loan.principal_cents),
                    None,
                    None,
                    None,
                    balance,
                    loan.disbursement_ref or "",
                ]
            )
        if loan.payments:
            splits = allocate_payment_splits(loan.schedule, loan.payments)
            for payment in loan.payments:
                if (date_from or date_to) and not _in_window(
                    payment.received_at, frm, to
                ):
                    continue
                split = splits[payment.id]
                rows.append(
                    [
                        vend,
                        prov,
                        loan_id,
                        name,
                        payment.received_at.date(),
                        "Payment",
                        _method_label(payment.method),
                        _dollars(payment.amount_cents),
                        _dollars(split["interest_cents"]),
                        _dollars(split["principal_cents"]),
                        0.0,
                        balance,
                        payment.external_ref or "",
                    ]
                )
    rows.sort(key=lambda r: (r[4], str(r[2])))
    return cols, rows


def _build_scheduled_payments(db, vendor_id, date_from, date_to):
    cols = [
        Col("Vendor"),
        Col("Provider"),
        Col("Loan ID"),
        Col("Full Name"),
        Col("Date", "date"),
        Col("Amount", "money", total=True),
        Col("Status"),
    ]
    today = datetime.now(timezone.utc).date()
    rows = []
    for loan, app, vendor, patient in _loan_rows(db, vendor_id):
        if loan.status in ("charged_off", "cancelled"):
            continue
        vend, prov = _vendor_names(vendor)
        name = _full_name(app, patient)
        loan_id = _loan_display_id(loan)

        # As-agreed installments still carrying an open balance. With no
        # explicit window: every open obligation, past due AND future,
        # matching TL's "expected payments" semantics.
        for item in loan.schedule:
            if item.status in ("paid", "waived"):
                continue
            if not (date_from is None and date_to is None) and not _date_in_window(
                item.due_date, date_from, date_to
            ):
                continue
            outstanding = item.total_cents - item.paid_cents
            if outstanding <= 0:
                continue
            rows.append(
                [
                    vend,
                    prov,
                    loan_id,
                    name,
                    item.due_date,
                    _dollars(outstanding),
                    _expected_status(item, today),
                ]
            )

        # Staff-added custom transactions (WS-F): one-off future payment
        # instructions layered on top of the plan — TL lists these among
        # scheduled transactions.
        for custom in loan.custom_transactions:
            if custom.status != "scheduled":
                continue
            if not (date_from is None and date_to is None) and not _date_in_window(
                custom.scheduled_date, date_from, date_to
            ):
                continue
            rows.append(
                [
                    vend,
                    prov,
                    loan_id,
                    name,
                    custom.scheduled_date,
                    _dollars(custom.amount_cents),
                    "Scheduled (Custom)",
                ]
            )
    rows.sort(key=lambda r: (r[4], str(r[2])))
    return cols, rows


# ---------------------------------------------------------------------------
# Registry + XLSX rendering
# ---------------------------------------------------------------------------

_REPORTS: dict[str, tuple[str, Callable]] = {
    "borrower-data": ("Borrower Data", _build_borrower_data),
    "loan-originations": ("New Loan Originations", _build_loan_originations),
    "payments-received": ("Successful Loan Payments", _build_payments_received),
    "payments-failed": ("Failed Loan Payments", _build_payments_failed),
    "payments-reversed": ("Reversed Loan Payments", _build_payments_reversed),
    "loan-transactions": ("Loan Transactions", _build_loan_transactions),
    "scheduled-payments": ("Scheduled Payments", _build_scheduled_payments),
}

REPORT_KEYS: tuple[str, ...] = tuple(_REPORTS)


def report_catalog() -> list[dict]:
    return [{"key": k, "title": t} for k, (t, _) in _REPORTS.items()]


_KIND_FORMATS = {
    "money": "#,##0.00",
    "percent": "0.00%",
    "int": "#,##0",
    "date": "yyyy-mm-dd",
    "datetime": "yyyy-mm-dd hh:mm",
}


def _render_xlsx(
    title: str,
    period_label: str,
    cols: Sequence[Col],
    rows: Iterable[Sequence],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    bold = Font(bold=True)
    for i, line in enumerate(COMPANY_HEADER, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = bold

    title_cell = ws.cell(row=7, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    if period_label:
        ws.cell(row=7, column=3, value=period_label)
    ws.cell(row=7, column=max(4, len(cols)), value=COMPANY_WEBSITE)

    header_row = 8
    for c, col in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c, value=col.header)
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(28, len(col.header) + 6))

    r = header_row
    n_rows = 0
    for row in rows:
        r += 1
        n_rows += 1
        for c, (col, value) in enumerate(zip(cols, row), start=1):
            # Strip tzinfo — xlsx cells store naive timestamps.
            if isinstance(value, datetime) and value.tzinfo is not None:
                value = value.astimezone(timezone.utc).replace(tzinfo=None)
            cell = ws.cell(row=r, column=c, value=value)
            fmt = _KIND_FORMATS.get(col.kind)
            if fmt:
                cell.number_format = fmt

    # Totals row over the money columns flagged total=True (TL SUBTOTAL parity).
    if n_rows:
        total_row = r + 1
        for c, col in enumerate(cols, start=1):
            if not col.total:
                continue
            letter = get_column_letter(c)
            cell = ws.cell(
                row=total_row,
                column=c,
                value=f"=SUBTOTAL(109,{letter}{header_row + 1}:{letter}{r})",
            )
            cell.number_format = _KIND_FORMATS["money"]
            cell.font = bold
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(cols))}{r}"
        )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_report(
    db: Session,
    key: str,
    *,
    vendor_id: Optional[UUID] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> ReportResult:
    """Build one of the seven TL-parity reports as an XLSX byte blob.

    ``vendor_id`` scopes to a single vendor (the clinic surface forces it);
    ``date_from``/``date_to`` are an inclusive date window on each report's
    natural time axis (origination date, payment effective date, due date).
    """
    if key not in _REPORTS:
        raise KeyError(f"Unknown report: {key!r}")
    title, builder = _REPORTS[key]
    cols, rows = builder(db, vendor_id, date_from, date_to)

    if date_from or date_to:
        period_label = f"{date_from.isoformat() if date_from else '…'} to {date_to.isoformat() if date_to else '…'}"
    else:
        period_label = ""

    content = _render_xlsx(title, period_label, cols, rows)
    stamp = datetime.now(timezone.utc).date().isoformat()
    filename = f"payspyre_{key.replace('-', '_')}_{stamp}.xlsx"
    return ReportResult(
        key=key, title=title, filename=filename, content=content, row_count=len(rows)
    )

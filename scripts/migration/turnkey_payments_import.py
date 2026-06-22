"""Run the Turnkey -> PaySpyre PAYMENT-HISTORY import end-to-end.

Loads historical loan payments out of a Turnkey payment-ledger export into the PaySpyre
``platform_loan_payments`` ledger so the AI training dataset has real performance signal.

DRY RUN by default (no writes). With --execute it persists the mapped payments via the
idempotent ledger-only persist step. Re-running is safe (dedups on the stable Turnkey
transaction id via the unique (loan_id, external_ref) index).

LEDGER-ONLY: imported payments are raw PlatformLoanPayment rows for analytics. They do
NOT run through record_payment, so the schedule and principal_balance of migrated loans
are never touched (re-applying historical cash on top of the already-snapshotted balance
would double-count — see app/services/migration/turnkey_payments.py).

SAFETY: real borrower data must never land on the dev-tools-enabled staging box, so the
script refuses a target URL that looks like staging unless --force is given.

Examples:
  # preview only
  python scripts/migration/turnkey_payments_import.py payments.xlsx
  # actually import into the target DB (loans must already be imported first)
  python scripts/migration/turnkey_payments_import.py payments.xlsx --execute \\
      --database-url postgresql+psycopg2://user:pw@host:5432/payspyre
"""
import argparse
import sys

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.services.migration.turnkey_payments import map_payments, persist_payments

# Sheet that holds the payment ledger and the header row within it. Adjust against the
# real export alongside the Col constants in turnkey_payments.py.
PAYMENTS_SHEET = "Payments"
HEADER_ROW = 1


def _rows_with_headers(sheet):
    """Yield dict rows keyed by the header row, matching the by-name Col constants."""
    it = sheet.iter_rows(min_row=HEADER_ROW, values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(it)]
    except StopIteration:
        return
    for values in it:
        if values is None or all(v is None for v in values):
            continue
        yield dict(zip(headers, values))


def main() -> None:
    ap = argparse.ArgumentParser(description="Turnkey -> PaySpyre payment-history import")
    ap.add_argument("excel", help="path to the Turnkey payment-ledger export .xlsx")
    ap.add_argument("--execute", action="store_true", help="actually write (default: dry run)")
    ap.add_argument("--database-url", help="target DB (required with --execute)")
    ap.add_argument("--force", action="store_true", help="override the staging-URL safety guard")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    sheet = wb[PAYMENTS_SHEET] if PAYMENTS_SHEET in wb.sheetnames else wb.active
    mapres = map_payments(_rows_with_headers(sheet))

    print(f"Parsed {mapres.total_rows} payment rows -> {len(mapres.payments)} valid "
          f"({len(mapres.invalid)} invalid, skipped).")
    print("  Turnkey -> PaySpyre mapping: Acct# -> loan.legacy_account_number, "
          "Amount -> amount_cents, Payment Date -> received_at, "
          "Transaction Id -> external_ref (namespaced 'turnkey:').")
    if mapres.payments:
        total = sum(p.amount_cents for p in mapres.payments)
        print(f"  total payment value: ${total/100:,.2f}")
    if mapres.invalid:
        print(f"  WARNING — {len(mapres.invalid)} invalid row(s):")
        for w in mapres.invalid[:10]:
            print(f"    - {w}")

    if not args.execute:
        print("\nDRY RUN — nothing written. Add --execute --database-url ... to persist.")
        return

    if not args.database_url:
        sys.exit("error: --database-url is required with --execute")
    if "staging" in args.database_url.lower() and not args.force:
        sys.exit(
            "REFUSING: target looks like staging. Real borrower data must not land on the "
            "dev-tools-enabled staging environment. Use --force only if you are certain."
        )

    engine = create_engine(args.database_url)
    db = sessionmaker(bind=engine)()
    try:
        res = persist_payments(db, mapres.payments, invalid_count=len(mapres.invalid), commit=True)
        print(f"\nPERSISTED (ledger-only — balances/schedules untouched):")
        print(f"  imported            {res.imported}")
        print(f"  skipped (duplicate) {res.skipped_duplicate}")
        print(f"  unmatched account   {res.unmatched_account}"
              + (f"  -> accts {res.unmatched_accts[:10]}" if res.unmatched_accts else ""))
        print(f"  invalid (mapping)   {res.invalid}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
